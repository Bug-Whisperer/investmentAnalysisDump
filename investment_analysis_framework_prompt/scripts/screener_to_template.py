#!/usr/bin/env python3
"""
Screener.in Export → Buffett Analysis Template Mapper
======================================================
Reads Screener.in's Excel export (.xlsx) or individual CSV table exports
and populates the Buffett Analysis Template.

PRIMARY MODE — Screener .xlsx export (recommended):
    python screener_to_template.py --screener ITC.xlsx --template Buffett_Analysis_Template.xlsx

    How to get the .xlsx:
      1. Go to https://www.screener.in/company/TICKER/consolidated/
      2. Click "Export to Excel" at the top of the page
      3. Save the downloaded .xlsx file
      4. Run this script

CSV MODE — individual table exports (fallback):
    python screener_to_template.py --template Buffett_Analysis_Template.xlsx \
        --pl profit_loss.csv --bs balance_sheet.csv --cf cash_flow.csv

    How to get the CSVs:
      1. On each financial table, click the small download icon (↓)
      2. Save each CSV file separately

Requirements:
    pip install openpyxl
"""

import argparse
import csv
import os
import sys
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: openpyxl")
    print("Install with: pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════
# TEMPLATE ROW MAP — must match Buffett_Analysis_Template.xlsx
# ══════════════════════════════════════════════════════════════

ROW_YEAR_LABELS = 11

# Income Statement (rows 12–28)
ROW_REVENUE       = 12
ROW_OP_PROFIT     = 13
ROW_DEPRECIATION  = 14
ROW_INTEREST      = 15
ROW_PBT           = 16
ROW_TAX           = 17
ROW_NET_PROFIT    = 18
ROW_EPS           = 19
ROW_DPS           = 20
# Bank-mode rows 22–28 (NII, Other Inc, Opex, Provisions, NIM)
ROW_BANK_INT_INC  = 22
ROW_BANK_INT_EXP  = 23
ROW_BANK_NII      = 24
ROW_BANK_OTHER_INC = 25
ROW_BANK_OPEX     = 26
ROW_BANK_PROVISIONS = 27
ROW_BANK_NIM      = 28

# Balance Sheet (rows 31–49)
ROW_TOTAL_ASSETS  = 31
ROW_TOTAL_LIAB    = 32
ROW_EQUITY        = 33
ROW_BORROWINGS    = 34
ROW_CASH          = 35
ROW_GOODWILL      = 36
ROW_RESERVES      = 37
ROW_BVPS          = 38
# Bank-mode rows 40–49
ROW_ADVANCES      = 40
ROW_DEPOSITS      = 41
ROW_CASA          = 42
ROW_GNPA_AMT      = 43
ROW_NNPA_AMT      = 44
ROW_GNPA_PCT      = 45
ROW_NNPA_PCT      = 46
ROW_PCR           = 47
ROW_CRAR          = 48
ROW_TIER1         = 49

# Cash Flow (rows 53–59)
ROW_CFO           = 53
ROW_CAPEX         = 54
ROW_CFI           = 55
ROW_CFF           = 56
ROW_SBC           = 57
ROW_BUYBACKS      = 58
ROW_DIV_PAID      = 59

# Additional inputs start around row 62
ROW_ADD_START     = 62


# ══════════════════════════════════════════════════════════════
# SCREENER .XLSX PARSER  (reads the "Data Sheet")
# ══════════════════════════════════════════════════════════════

def parse_screener_xlsx(filepath: str) -> dict:
    """Parse Screener.in's native .xlsx export via its 'Data Sheet'.

    Returns a structured dict with all sections, metadata, and values.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if 'Data Sheet' not in wb.sheetnames:
        print("  ERROR: No 'Data Sheet' found. Is this a Screener.in export?")
        print(f"  Sheets found: {wb.sheetnames}")
        return {}

    ws = wb['Data Sheet']
    max_row = ws.max_row
    max_col = ws.max_column

    # Helper: read a row as (label, [values...])
    def read_row(r):
        label = ws.cell(row=r, column=1).value
        vals = []
        for c in range(2, max_col + 1):
            v = ws.cell(row=r, column=c).value
            vals.append(v)
        return str(label).strip() if label else '', vals

    # ── META ──
    meta = {}
    company_name = ''
    row1_label, row1_vals = read_row(1)
    if row1_vals and row1_vals[0]:
        company_name = str(row1_vals[0]).strip()

    for r in range(5, 15):
        label, vals = read_row(r)
        if not label:
            continue
        ll = label.lower()
        val = vals[0] if vals else None
        if 'number of shares' in ll:
            meta['shares_raw'] = val  # might be None if it's a formula
        elif 'face value' in ll:
            meta['face_value'] = val
        elif 'current price' in ll:
            meta['current_price'] = val
        elif 'market cap' in ll:
            meta['market_cap'] = val

    # ── Find section boundaries by scanning for known headers ──
    sections = {}  # name -> start_row
    for r in range(1, max_row + 1):
        label, _ = read_row(r)
        ll = label.lower().strip().rstrip(':')
        if ll in ('profit & loss', 'profit and loss'):
            sections['pl'] = r
        elif ll == 'quarters':
            sections['quarters'] = r
        elif ll == 'balance sheet':
            sections['bs'] = r
        elif ll in ('cash flow', 'cash flow:'):
            sections['cf'] = r
        elif ll in ('price', 'price:'):
            sections['price'] = r
        elif ll in ('derived', 'derived:'):
            sections['derived'] = r

    def parse_section(start_row: int, end_row: int) -> dict:
        """Parse a section into {metric_name: [values], '_headers': [year_labels]}."""
        result = {'_headers': []}
        for r in range(start_row + 1, end_row):
            label, vals = read_row(r)
            if not label:
                continue
            ll = label.lower().strip()

            # The "Report Date" row contains year labels
            if 'report date' in ll:
                year_labels = []
                for v in vals:
                    if v is None:
                        continue
                    if isinstance(v, datetime):
                        year_labels.append(f"Mar {v.year}")
                    elif isinstance(v, str) and v.strip():
                        # Try to extract year from string like "2024-03-31 00:00:00"
                        m = re.search(r'(\d{4})', v)
                        if m:
                            year_labels.append(f"Mar {m.group(1)}")
                        else:
                            year_labels.append(v.strip())
                    else:
                        try:
                            year_labels.append(str(v))
                        except:
                            pass
                result['_headers'] = year_labels
                continue

            # Data row — extract numeric values
            numeric = []
            for v in vals:
                if v is None:
                    numeric.append(None)
                elif isinstance(v, (int, float)):
                    numeric.append(float(v))
                elif isinstance(v, str):
                    cleaned = v.strip().replace(',', '').replace('%', '')
                    if cleaned in ('', '-'):
                        numeric.append(None)
                    else:
                        try:
                            numeric.append(float(cleaned))
                        except ValueError:
                            numeric.append(None)
                else:
                    numeric.append(None)

            # Keep all values including trailing Nones — all metrics must have
            # the same length so write_row right-aligns them to the same columns.
            # (Trimming independently would shift shorter lists rightward.)
            if any(v is not None for v in numeric):
                result[label] = numeric

        return result

    # Determine section end boundaries
    all_starts = sorted(sections.values())

    def section_end(start):
        for s in all_starts:
            if s > start:
                return s
        return max_row + 1

    parsed = {
        'company_name': company_name,
        'meta': meta,
        'pl': {},
        'bs': {},
        'cf': {},
        'price': {},
        'derived': {},
        'quarters': {},
    }

    for name, start in sections.items():
        end = section_end(start)
        parsed[name] = parse_section(start, end)

    # Extract shares from derived section or BS
    if 'Adjusted Equity Shares in Cr' in parsed.get('derived', {}):
        shares_cr = parsed['derived']['Adjusted Equity Shares in Cr']
        if shares_cr:
            meta['shares_cr'] = shares_cr[-1]  # latest
    elif 'No. of Equity Shares' in parsed.get('bs', {}):
        shares_abs = parsed['bs']['No. of Equity Shares']
        if shares_abs:
            latest = shares_abs[-1]
            if latest and latest > 1e6:
                meta['shares_cr'] = latest / 1e7  # absolute → crores

    wb.close()

    n_pl = len([k for k in parsed['pl'] if not k.startswith('_')])
    n_bs = len([k for k in parsed['bs'] if not k.startswith('_')])
    n_cf = len([k for k in parsed['cf'] if not k.startswith('_')])
    print(f"  Company: {company_name}")
    print(f"  CMP: ₹{meta.get('current_price', 'N/A')}")
    print(f"  Market Cap: ₹{meta.get('market_cap', 'N/A')} Cr")
    print(f"  Shares: {meta.get('shares_cr', 'N/A')} Cr")
    print(f"  P&L metrics: {n_pl}, BS metrics: {n_bs}, CF metrics: {n_cf}")
    print(f"  Years: {parsed['pl'].get('_headers', [])}")

    return parsed


# ══════════════════════════════════════════════════════════════
# SCREENER CSV PARSER  (fallback for individual table exports)
# ══════════════════════════════════════════════════════════════

def read_screener_csv(filepath: str) -> dict:
    """Read a Screener.in CSV export into the same format as parse_section."""
    filepath = Path(filepath)
    if not filepath.exists():
        print(f"  ERROR: File not found: {filepath}")
        return {}

    rows = []
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(filepath, newline='', encoding=enc) as f:
                rows = list(csv.reader(f))
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if not rows or len(rows) < 2:
        print(f"  ERROR: CSV is empty or has < 2 rows: {filepath}")
        return {}

    raw_headers = rows[0]
    headers = [h.strip() for h in raw_headers[1:] if h.strip()]

    result = {'_headers': headers}
    for row in rows[1:]:
        if not row or not row[0].strip():
            continue
        metric_name = row[0].strip()
        if metric_name in ('+', '-', ''):
            continue
        values = []
        for cell in row[1:1 + len(headers)]:
            cell = cell.strip().replace(',', '').replace('%', '')
            if cell == '' or cell == '-':
                values.append(None)
            else:
                try:
                    values.append(float(cell))
                except ValueError:
                    values.append(None)
        result[metric_name] = values

    print(f"  Parsed {filepath.name}: {len(headers)} periods, "
          f"{len(result) - 1} metrics")
    return result


def detect_csv_section(parsed: dict) -> str:
    """Auto-detect whether a CSV is P&L, BS, CF, or Ratios."""
    if not parsed:
        return 'unknown'
    keys_lower = {k.lower() for k in parsed if not k.startswith('_')}
    if any('sales' in k or 'operating profit' in k for k in keys_lower):
        return 'pl'
    if any('total assets' in k or 'borrowings' in k or 'reserves' in k
           for k in keys_lower):
        return 'bs'
    if any('cash from operating' in k or 'operating activity' in k
           for k in keys_lower):
        return 'cf'
    if any('return on equity' in k or 'debtor days' in k for k in keys_lower):
        return 'ratios'
    return 'unknown'


# ══════════════════════════════════════════════════════════════
# METRIC LOOKUP
# ══════════════════════════════════════════════════════════════

def find_values(section: dict, keywords: list, exclude: list = None) -> list:
    """Find a metric's value list by substring-matching against section keys.

    Args:
        section: Dict from parse_section / read_screener_csv.
        keywords: Substrings to match (case-insensitive, first match wins).
        exclude: Substrings that disqualify a match.
    """
    if not section:
        return []
    exclude_lower = [e.lower() for e in (exclude or [])]

    for key in section:
        if key.startswith('_'):
            continue
        kl = key.lower().strip()
        if any(ex in kl for ex in exclude_lower):
            continue
        for kw in keywords:
            if kw.lower() in kl:
                return section[key]
    return []


# ══════════════════════════════════════════════════════════════
# TEMPLATE POPULATION
# ══════════════════════════════════════════════════════════════

def write_row(ws, row_num: int, values: list, max_cols: int = 11):
    """Write values into template row, right-aligned (latest in col L=12).

    Takes up to max_cols values from the END of the list.
    Column B=2 is oldest, column L=12 is latest/TTM.
    """
    if not values:
        return
    n = min(len(values), max_cols)
    tail = values[-n:]
    for i, val in enumerate(tail):
        col = (max_cols - n) + i + 2  # right-align into B..L
        if 2 <= col <= 12 and val is not None:
            ws.cell(row=row_num, column=col, value=val)


def populate_from_screener_xlsx(ws, parsed: dict):
    """Map parsed Screener .xlsx data into the template's input sheet."""

    meta = parsed.get('meta', {})
    pl   = parsed.get('pl', {})
    bs   = parsed.get('bs', {})
    cf   = parsed.get('cf', {})
    price_section = parsed.get('price', {})
    derived = parsed.get('derived', {})
    company = parsed.get('company_name', '')

    filled = 0

    # ── Company header ──
    ws['B2'] = company
    ws['B3'] = company  # ticker slot — user can rename
    if meta.get('current_price'):
        ws['B4'] = meta['current_price']
    if meta.get('shares_cr'):
        ws['B5'] = round(meta['shares_cr'], 2)

    # ── Year labels ──
    headers = pl.get('_headers') or bs.get('_headers') or cf.get('_headers') or []
    n = min(len(headers), 11)
    for i in range(n):
        col = (11 - n) + i + 2
        if 2 <= col <= 12:
            ws.cell(row=ROW_YEAR_LABELS, column=col, value=headers[-(n - i)])

    # ── Profit & Loss ──
    pl_map = [
        (ROW_REVENUE,      ['Sales', 'Revenue', 'Total Income from Operations'], None),
        (ROW_OP_PROFIT,    ['Operating Profit'], None),
        (ROW_DEPRECIATION, ['Depreciation'], None),
        (ROW_INTEREST,     ['Interest'], ['net interest income']),
        (ROW_PBT,          ['Profit before tax'], None),
        (ROW_TAX,          ['Tax'], ['deferred tax', 'tax %', 'profit before']),
        (ROW_NET_PROFIT,   ['Net profit', 'Net Profit'], None),
        # Bank P&L rows
        (ROW_BANK_INT_INC,    ['Interest Earned', 'Interest Income'], None),
        (ROW_BANK_INT_EXP,    ['Interest Expended'], None),
        (ROW_BANK_NII,        ['Net Interest Income'], None),
        (ROW_BANK_OTHER_INC,  ['Other Income'], ['total']),
        (ROW_BANK_OPEX,       ['Operating Expenses'], None),
        (ROW_BANK_PROVISIONS, ['Provisions and Contingencies', 'Provisions'], ['coverage']),
    ]

    for row_num, keywords, exclusions in pl_map:
        vals = find_values(pl, keywords, exclude=exclusions)
        if vals:
            write_row(ws, row_num, vals)
            filled += 1

    # Operating Profit — Screener's Data Sheet often doesn't have this as a
    # named row; it's computed in the display sheet.  Compute from:
    # Sales - (Raw Material + Change in Inventory + Power + Other Mfr +
    #          Employee + Selling & Admin + Other Expenses)
    if not find_values(pl, ['Operating Profit']):
        sales = find_values(pl, ['Sales'])
        expense_keys = [
            'Raw Material Cost', 'Change in Inventory', 'Power and Fuel',
            'Other Mfr. Exp', 'Employee Cost', 'Selling and admin',
            'Other Expenses',
        ]
        if sales:
            n = len(sales)
            total_expenses = [0.0] * n
            found_any_expense = False
            for ek in expense_keys:
                ev = find_values(pl, [ek])
                if ev:
                    found_any_expense = True
                    for i in range(min(len(ev), n)):
                        if ev[i] is not None:
                            total_expenses[i] += ev[i]
            if found_any_expense:
                op_profit = []
                for i in range(n):
                    if sales[i] is not None:
                        op_profit.append(round(sales[i] - total_expenses[i], 2))
                    else:
                        op_profit.append(None)
                write_row(ws, ROW_OP_PROFIT, op_profit)
                filled += 1
                print("  → Operating Profit computed from Sales - Expenses")

    # EPS — compute from Net Profit / Shares if not directly available
    eps_vals = find_values(pl, ['EPS'])
    if not eps_vals:
        np_vals = find_values(pl, ['Net profit', 'Net Profit'])
        shares_series = find_values(derived, ['Adjusted Equity Shares in Cr'])
        if not shares_series:
            shares_series = find_values(bs, ['No. of Equity Shares'])
            if shares_series:
                shares_series = [s / 1e7 if s and s > 1e6 else s for s in shares_series]
        if np_vals and shares_series:
            n = min(len(np_vals), len(shares_series))
            eps_vals = []
            for i in range(n):
                np_v = np_vals[i]
                sh_v = shares_series[i]
                if np_v is not None and sh_v and sh_v > 0:
                    eps_vals.append(round(np_v / sh_v, 2))
                else:
                    eps_vals.append(None)
            print("  → EPS computed from Net Profit / Shares")
    if eps_vals:
        write_row(ws, ROW_EPS, eps_vals)
        filled += 1

    # DPS — compute from Dividend Amount / Shares
    div_amt = find_values(pl, ['Dividend Amount'])
    if div_amt:
        shares_series = find_values(derived, ['Adjusted Equity Shares in Cr'])
        if not shares_series:
            shares_series = find_values(bs, ['No. of Equity Shares'])
            if shares_series:
                shares_series = [s / 1e7 if s and s > 1e6 else s for s in shares_series]
        if shares_series:
            n = min(len(div_amt), len(shares_series))
            dps_vals = []
            for i in range(n):
                d = div_amt[i]
                s = shares_series[i]
                if d is not None and s and s > 0:
                    dps_vals.append(round(d / s, 2))
                else:
                    dps_vals.append(None)
            write_row(ws, ROW_DPS, dps_vals)
            filled += 1
            print("  → DPS computed from Dividend Amount / Shares")

    # ── Balance Sheet ──
    bs_map = [
        (ROW_BORROWINGS, ['Borrowings', 'Total Debt'], None),
        (ROW_RESERVES,   ['Reserves'], None),
        (ROW_GOODWILL,   ['Goodwill', 'Intangible'], None),
        # Bank-mode
        (ROW_ADVANCES,   ['Total Advances', 'Advances'], ['growth']),
        (ROW_DEPOSITS,   ['Total Deposits', 'Deposits'], ['growth', 'casa', 'cost']),
        (ROW_CASA,       ['CASA'], None),
    ]

    for row_num, keywords, exclusions in bs_map:
        vals = find_values(bs, keywords, exclude=exclusions)
        if vals:
            write_row(ws, row_num, vals)
            filled += 1

    # Total Assets — Screener uses "Total" which appears twice in BS.
    # The SECOND "Total" is assets-side. Use it directly.
    total_vals = []
    total_count = 0
    for key in bs:
        if key.startswith('_'):
            continue
        if key.strip().lower() == 'total':
            total_count += 1
            total_vals_candidate = bs[key]
            if total_count == 1:
                # First "Total" = Liabilities + Equity side
                total_vals = total_vals_candidate
            if total_count == 2:
                # Second "Total" = Assets side (should be same number)
                break

    if total_vals:
        write_row(ws, ROW_TOTAL_ASSETS, total_vals)
        filled += 1
        # Total Liabilities = Total Assets - Equity
        # We'll compute equity first, then derive liabilities

    # Equity = Equity Share Capital + Reserves
    sc_vals = find_values(bs, ['Equity Share Capital', 'Share Capital'])
    res_vals = find_values(bs, ['Reserves'])
    if sc_vals and res_vals:
        n = min(len(sc_vals), len(res_vals))
        equity_vals = [(sc_vals[i] or 0) + (res_vals[i] or 0) for i in range(n)]
        write_row(ws, ROW_EQUITY, equity_vals)
        filled += 1

        # Total Liabilities = Total - Equity
        if total_vals:
            n2 = min(len(total_vals), len(equity_vals))
            liab_vals = [(total_vals[i] or 0) - (equity_vals[i] or 0)
                         for i in range(n2)]
            write_row(ws, ROW_TOTAL_LIAB, liab_vals)
            filled += 1
    else:
        # Fallback: try "Total Shareholders Funds"
        eq_direct = find_values(bs, ['Total Shareholders Funds', 'Shareholders Fund'],
                                exclude=['equity share capital'])
        if eq_direct:
            write_row(ws, ROW_EQUITY, eq_direct)
            filled += 1

    # Cash & Bank
    cash_vals = find_values(bs, ['Cash & Bank', 'Cash Equivalents', 'Cash'])
    if cash_vals:
        write_row(ws, ROW_CASH, cash_vals)
        filled += 1

    # BVPS = Equity / Shares
    if sc_vals and res_vals:
        shares_series = find_values(derived, ['Adjusted Equity Shares in Cr'])
        if not shares_series:
            shares_series = find_values(bs, ['No. of Equity Shares'])
            if shares_series:
                shares_series = [s / 1e7 if s and s > 1e6 else s
                                 for s in shares_series]
        if shares_series:
            n = min(len(equity_vals), len(shares_series))
            bvps = []
            for i in range(n):
                eq = equity_vals[i] if i < len(equity_vals) else 0
                sh = shares_series[i]
                if eq and sh and sh > 0:
                    bvps.append(round(eq / sh, 2))
                else:
                    bvps.append(None)
            write_row(ws, ROW_BVPS, bvps)
            filled += 1
            print("  → BVPS computed from (Share Capital + Reserves) / Shares")

    # ── Cash Flow ──
    cf_map = [
        (ROW_CFO,   ['Cash from Operating Activity'], None),
        (ROW_CFI,   ['Cash from Investing Activity'], None),
        (ROW_CFF,   ['Cash from Financing Activity'], None),
    ]
    for row_num, keywords, exclusions in cf_map:
        vals = find_values(cf, keywords, exclude=exclusions)
        if vals:
            write_row(ws, row_num, vals)
            filled += 1

    # CapEx — Screener doesn't break this out in Cash Flow section.
    # It's embedded in CFI. Check if "Fixed Assets Purchased" exists.
    capex = find_values(cf, ['Fixed Assets Purchased', 'Purchase of Fixed Assets',
                             'Capital Expenditure'])
    if capex:
        write_row(ws, ROW_CAPEX, capex)
        filled += 1

    # Dividends Paid (from P&L Dividend Amount, negated for CF convention)
    if div_amt:
        div_paid = [(-abs(d) if d else None) for d in div_amt]
        write_row(ws, ROW_DIV_PAID, div_paid)
        filled += 1

    # ── Additional inputs from meta / ratios ──
    if meta.get('current_price'):
        # Current PE = Price / latest EPS
        if eps_vals:
            latest_eps = next((v for v in reversed(eps_vals) if v and v > 0), None)
            if latest_eps:
                ws.cell(row=ROW_ADD_START + 3, column=2,
                         value=round(meta['current_price'] / latest_eps, 2))
        # Current P/B
        bvps_latest = ws.cell(row=ROW_BVPS, column=12).value
        if not bvps_latest:
            bvps_latest = ws.cell(row=ROW_BVPS, column=11).value
        if bvps_latest and bvps_latest > 0:
            ws.cell(row=ROW_ADD_START + 4, column=2,
                     value=round(meta['current_price'] / bvps_latest, 2))

    # ROE from BS (if available)
    roe_vals = find_values(bs, ['Return on Equity'])
    if roe_vals:
        # Store latest ROE for reference
        latest_roe = next((v for v in reversed(roe_vals) if v is not None), None)
        if latest_roe:
            print(f"  → Latest ROE from Screener: {latest_roe}%")

    print(f"\n  Populated {filled} metric rows into the template.")
    return filled


def populate_from_csvs(ws, pl_data, bs_data, cf_data, ratios_data,
                       ticker, price, shares):
    """Map CSV-parsed data into template. (Legacy CSV mode.)"""

    filled = 0
    ws['B2'] = ticker
    ws['B3'] = ticker
    if price:
        ws['B4'] = price
    if shares:
        ws['B5'] = shares

    # Year labels
    all_parsed = [d for d in (pl_data, bs_data, cf_data, ratios_data) if d]
    if all_parsed:
        best = max(all_parsed, key=lambda d: len(d.get('_headers', [])))
        headers = best['_headers']
        n = min(len(headers), 11)
        for i in range(n):
            col = (11 - n) + i + 2
            if 2 <= col <= 12:
                ws.cell(row=ROW_YEAR_LABELS, column=col, value=headers[-(n - i)])

    # P&L
    if pl_data:
        for row_num, kw, ex in [
            (ROW_REVENUE,      ['Sales', 'Revenue'], None),
            (ROW_OP_PROFIT,    ['Operating Profit'], None),
            (ROW_DEPRECIATION, ['Depreciation'], None),
            (ROW_INTEREST,     ['Interest'], ['net interest income']),
            (ROW_PBT,          ['Profit before tax'], None),
            (ROW_TAX,          ['Tax'], ['deferred tax', 'tax %', 'profit before']),
            (ROW_NET_PROFIT,   ['Net profit', 'Net Profit'], None),
            (ROW_EPS,          ['EPS in Rs', 'EPS'], ['growth']),
            (ROW_DPS,          ['Dividend Per Share'], ['payout']),
        ]:
            vals = find_values(pl_data, kw, exclude=ex)
            if vals:
                write_row(ws, row_num, vals)
                filled += 1

    # BS
    if bs_data:
        for row_num, kw, ex in [
            (ROW_TOTAL_ASSETS, ['Total Assets'], None),
            (ROW_TOTAL_LIAB,   ['Total Liabilities'], None),
            (ROW_BORROWINGS,   ['Borrowings', 'Total Debt'], None),
            (ROW_CASH,         ['Cash', 'Cash Equivalents'], None),
            (ROW_GOODWILL,     ['Goodwill', 'Intangible'], None),
            (ROW_RESERVES,     ['Reserves'], None),
        ]:
            vals = find_values(bs_data, kw, exclude=ex)
            if vals:
                write_row(ws, row_num, vals)
                filled += 1

        # Equity
        eq = find_values(bs_data,
                         ['Total Shareholders Funds', 'Shareholders Fund'],
                         exclude=['equity share capital'])
        if eq:
            write_row(ws, ROW_EQUITY, eq)
            filled += 1
        else:
            sc = find_values(bs_data, ['Equity Share Capital', 'Share Capital'])
            rs = find_values(bs_data, ['Reserves'])
            if sc and rs:
                n = min(len(sc), len(rs))
                combined = [(sc[i] or 0) + (rs[i] or 0) for i in range(n)]
                write_row(ws, ROW_EQUITY, combined)
                filled += 1
                print("  → Equity = Share Capital + Reserves (fallback)")

    # CF
    if cf_data:
        for row_num, kw, ex in [
            (ROW_CFO,   ['Cash from Operating', 'Operating Activity'], None),
            (ROW_CAPEX, ['Fixed Assets Purchased', 'Purchase of Fixed Assets',
                         'Capital Expenditure'], None),
            (ROW_CFI,   ['Cash from Investing', 'Investing Activity'], None),
            (ROW_CFF,   ['Cash from Financing', 'Financing Activity'], None),
        ]:
            vals = find_values(cf_data, kw, exclude=ex)
            if vals:
                write_row(ws, row_num, vals)
                filled += 1

    # Ratios
    if ratios_data:
        bvps = find_values(ratios_data, ['Book Value', 'BVPS'])
        if bvps:
            write_row(ws, ROW_BVPS, bvps)
            filled += 1

    print(f"\n  Populated {filled} metric rows into the template.")
    return filled


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Map Screener.in exports into the Buffett Analysis Template.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # From Screener .xlsx export (RECOMMENDED):
  python screener_to_template.py --screener ITC.xlsx \\
      --template Buffett_Analysis_Template.xlsx

  # Override price/shares if Screener data is stale:
  python screener_to_template.py --screener ITC.xlsx \\
      --template Buffett_Analysis_Template.xlsx --price 450 --shares 12.51

  # From individual CSVs (fallback):
  python screener_to_template.py --template Buffett_Analysis_Template.xlsx \\
      --pl profit_loss.csv --bs balance_sheet.csv --cf cash_flow.csv

  # Auto-detect CSV sections:
  python screener_to_template.py --template Buffett_Analysis_Template.xlsx \\
      --auto pl.csv bs.csv cf.csv ratios.csv --ticker ITC
        """)

    parser.add_argument('--screener', default=None,
                        help='Path to Screener.in .xlsx export (the single file '
                             'you get from "Export to Excel")')
    parser.add_argument('--template', required=True,
                        help='Path to Buffett_Analysis_Template.xlsx')
    parser.add_argument('--pl', default=None, help='P&L CSV (CSV mode)')
    parser.add_argument('--bs', default=None, help='Balance Sheet CSV (CSV mode)')
    parser.add_argument('--cf', default=None, help='Cash Flow CSV (CSV mode)')
    parser.add_argument('--ratios', default=None, help='Ratios CSV (CSV mode)')
    parser.add_argument('--auto', nargs='+', default=None,
                        help='Auto-detect CSVs (CSV mode)')
    parser.add_argument('--ticker', default=None, help='Ticker override')
    parser.add_argument('--price', type=float, default=None,
                        help='CMP override (₹)')
    parser.add_argument('--shares', type=float, default=None,
                        help='Shares override (in Crores)')
    parser.add_argument('--output', default=None,
                        help='Output filename (default: {TICKER}_Buffett_Analysis.xlsx)')

    args = parser.parse_args()

    print(f"\n{'═' * 60}")
    print(f"  SCREENER → BUFFETT TEMPLATE MAPPER")
    print(f"{'═' * 60}\n")

    if not os.path.exists(args.template):
        print(f"ERROR: Template not found: {args.template}")
        sys.exit(1)

    # Load template
    print(f"Loading template: {args.template}")
    wb = openpyxl.load_workbook(args.template)
    ws = wb['1. Inputs']

    ticker = args.ticker or 'COMPANY'

    if args.screener:
        # ── XLSX MODE ──
        print(f"\nParsing Screener export: {args.screener}\n")
        parsed = parse_screener_xlsx(args.screener)
        if not parsed or not parsed.get('pl'):
            print("ERROR: Failed to parse Screener .xlsx export.")
            sys.exit(1)

        ticker = args.ticker or parsed.get('company_name', 'COMPANY')
        populate_from_screener_xlsx(ws, parsed)

        # Apply overrides
        if args.price:
            ws['B4'] = args.price
        if args.shares:
            ws['B5'] = args.shares
        if args.ticker:
            ws['B3'] = args.ticker

    elif args.auto or args.pl:
        # ── CSV MODE ──
        pl_d, bs_d, cf_d, rat_d = None, None, None, None

        if args.auto:
            print("Mode: Auto-detect CSVs\n")
            for path in args.auto:
                parsed = read_screener_csv(path)
                section = detect_csv_section(parsed)
                print(f"  {path} → {section.upper()}")
                if section == 'pl':
                    pl_d = parsed
                elif section == 'bs':
                    bs_d = parsed
                elif section == 'cf':
                    cf_d = parsed
                elif section == 'ratios':
                    rat_d = parsed
        else:
            print("Mode: Explicit CSV assignment\n")
            if args.pl:
                pl_d = read_screener_csv(args.pl)
            if args.bs:
                bs_d = read_screener_csv(args.bs)
            if args.cf:
                cf_d = read_screener_csv(args.cf)
            if args.ratios:
                rat_d = read_screener_csv(args.ratios)

        if not any([pl_d, bs_d, cf_d, rat_d]):
            print("ERROR: No valid data found.")
            sys.exit(1)

        populate_from_csvs(ws, pl_d, bs_d, cf_d, rat_d,
                           ticker, args.price or 0, args.shares or 0)
    else:
        print("ERROR: Provide either --screener (xlsx) or --pl/--auto (CSVs).")
        parser.print_help()
        sys.exit(1)

    output = args.output or f"{ticker.replace(' ', '_')}_Buffett_Analysis.xlsx"
    wb.save(output)

    print(f"\n{'═' * 60}")
    print(f"  DONE! Saved: {output}")
    print(f"{'═' * 60}")
    print(f"\nNext steps:")
    print(f"  1. Open {output} in Excel")
    print(f"  2. Verify auto-filled data on Sheet '1. Inputs'")
    print(f"  3. Fill remaining yellow cells (assumptions, qualitative inputs)")
    print(f"  4. Review computed sheets (2–6, 16)")
    print(f"  5. Enter DCF assumptions on Sheet 9")
    print(f"  6. Score the Final Scorecard on Sheet 17")


if __name__ == '__main__':
    main()
