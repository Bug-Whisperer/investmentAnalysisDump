#!/usr/bin/env python3
"""
Buffett Analysis Data Fetcher & Excel Builder
==============================================
Fetches financial data from free sources (no API keys needed) and 
populates the Buffett Analysis Excel template.

Usage:
    python fetch_and_build_analysis.py TICKER
    
Example:
    python fetch_and_build_analysis.py HDFCBANK
    python fetch_and_build_analysis.py TCS
    python fetch_and_build_analysis.py ITC

Requirements:
    pip install requests beautifulsoup4 openpyxl lxml

Data Sources (all free, no API key):
    - Screener.in (financial statements, ratios)
    - BSE/NSE websites (current price)

Output:
    Creates: {TICKER}_Buffett_Analysis.xlsx
"""

import sys
import re
import json
import time
import os
from pathlib import Path

try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install requests beautifulsoup4 openpyxl lxml")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════
# DATA FETCHING
# ══════════════════════════════════════════════════════════════

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
}


def fetch_screener_data(ticker: str) -> dict:
    """Fetch all financial data from Screener.in consolidated page."""
    url = f"https://www.screener.in/company/{ticker}/consolidated/"
    print(f"Fetching: {url}")
    
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        if resp.status_code == 404:
            # Try standalone
            url = f"https://www.screener.in/company/{ticker}/"
            print(f"Consolidated not found, trying: {url}")
            resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"Error fetching Screener data: {e}")
        return {}
    
    soup = BeautifulSoup(resp.text, 'lxml')
    data = {
        'company_name': '',
        'bse_code': '',
        'sector': '',
        'current_price': 0,
        'market_cap': 0,
        'pe_ratio': 0,
        'pb_ratio': 0,
        'face_value': 0,
        'shares': 0,
        'income_statement': {},
        'balance_sheet': {},
        'cash_flow': {},
        'ratios': {},
        'quarterly': {},
    }
    
    # Company name
    h1 = soup.find('h1')
    if h1:
        data['company_name'] = h1.get_text(strip=True)
    
    # Key ratios from top section
    top_ratios = soup.find('div', id='top-ratios')
    if not top_ratios:
        top_ratios = soup.find('ul', id='top-ratios')
    
    if top_ratios:
        spans = top_ratios.find_all('span', class_='number')
        names = top_ratios.find_all('span', class_='name')
        for name_el, val_el in zip(names, spans):
            name = name_el.get_text(strip=True).lower()
            val_text = val_el.get_text(strip=True).replace(',', '').replace('₹', '').strip()
            try:
                val = float(val_text)
            except (ValueError, TypeError):
                val = 0
            
            if 'market cap' in name:
                data['market_cap'] = val
            elif 'current price' in name:
                data['current_price'] = val
            elif 'stock p/e' in name or name == 'p/e':
                data['pe_ratio'] = val
            elif 'book value' in name:
                data['book_value'] = val
            elif 'face value' in name:
                data['face_value'] = val
    
    # Also try the company info list
    li_items = soup.find_all('li', class_='flex flex-space-between')
    for li in li_items:
        name_span = li.find('span', class_='name')
        val_span = li.find('span', class_='number')
        if name_span and val_span:
            name = name_span.get_text(strip=True).lower()
            val_text = val_span.get_text(strip=True).replace(',', '').replace('₹', '').replace('%', '').strip()
            try:
                val = float(val_text)
            except:
                val = 0
            if 'market cap' in name:
                data['market_cap'] = val
            elif 'current price' in name:
                data['current_price'] = val
            elif 'stock p/e' in name:
                data['pe_ratio'] = val
            elif 'book value' in name:
                data['book_value'] = val
            elif 'face value' in name:
                data['face_value'] = val
    
    # Compute shares from market cap and price
    if data['current_price'] > 0 and data['market_cap'] > 0:
        data['shares'] = data['market_cap'] / data['current_price']  # in Cr
    
    # Parse financial tables
    sections = soup.find_all('section')
    for section in sections:
        section_id = section.get('id', '')
        h2 = section.find('h2')
        if not h2:
            continue
        title = h2.get_text(strip=True).lower()
        
        table = section.find('table', class_='data-table')
        if not table:
            continue
        
        parsed = parse_screener_table(table)
        
        if 'profit' in title and 'loss' in title:
            data['income_statement'] = parsed
        elif 'balance' in title:
            data['balance_sheet'] = parsed
        elif 'cash' in title and 'flow' in title:
            data['cash_flow'] = parsed
        elif 'ratio' in title:
            data['ratios'] = parsed
        elif 'quarter' in title:
            data['quarterly'] = parsed
    
    return data


def parse_screener_table(table) -> dict:
    """Parse a Screener.in financial data table into structured data."""
    result = {}
    
    # Get headers (years)
    thead = table.find('thead')
    headers = []
    if thead:
        ths = thead.find_all('th')
        for th in ths:
            text = th.get_text(strip=True)
            headers.append(text)
    
    # Get rows
    tbody = table.find('tbody')
    if not tbody:
        tbody = table
    
    rows = tbody.find_all('tr')
    for row in rows:
        cells = row.find_all(['td', 'th'])
        if len(cells) < 2:
            continue
        
        # First cell is the metric name
        metric = cells[0].get_text(strip=True)
        if not metric or metric == '+':
            continue
        
        values = []
        for cell in cells[1:]:
            text = cell.get_text(strip=True).replace(',', '').replace('%', '').strip()
            try:
                values.append(float(text))
            except (ValueError, TypeError):
                values.append(None)
        
        result[metric] = {
            'headers': headers[1:] if headers else [],
            'values': values
        }
    
    return result


def find_metric(data_section: dict, keywords: list, exclude: list = None) -> tuple:
    """Find a metric in parsed data by matching keywords.
    
    Args:
        data_section: Parsed Screener table dict.
        keywords: Substrings to match against metric names.
        exclude: Substrings that disqualify a match (e.g., exclude 'Equity Share Capital'
                 when searching for total shareholders' equity).
    """
    if not data_section:
        return [], []
    
    exclude = [e.lower() for e in (exclude or [])]
    
    for key in data_section:
        key_lower = key.lower().strip()
        if any(ex in key_lower for ex in exclude):
            continue
        for kw in keywords:
            if kw.lower() in key_lower:
                entry = data_section[key]
                return entry.get('headers', []), entry.get('values', [])
    return [], []


# ══════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════

# Style constants
BLUE_FONT = Font(name='Arial', color='0000FF', size=10)
BLACK_FONT = Font(name='Arial', color='000000', size=10)
BLACK_BOLD = Font(name='Arial', color='000000', size=10, bold=True)
GREEN_FONT = Font(name='Arial', color='008000', size=10)
WHITE_BOLD = Font(name='Arial', color='FFFFFF', size=10, bold=True)
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
SUB_HEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
INPUT_FILL = PatternFill('solid', fgColor='FFFFCC')
CALC_FILL = PatternFill('solid', fgColor='E2EFDA')
LIGHT_BLUE_FILL = PatternFill('solid', fgColor='DAEEF3')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def build_prefilled_excel(ticker: str, data: dict, output_path: str):
    """Build a pre-filled Buffett Analysis Excel from fetched data."""
    
    TEMPLATE_NAME = 'Buffett_Analysis_Template.xlsx'
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Search for template in multiple likely locations:
    #   1. Same directory as the script
    #   2. Parent directory of the script (common layout: project_root/scripts/)
    #   3. Current working directory
    search_paths = [
        os.path.join(script_dir, TEMPLATE_NAME),
        os.path.join(os.path.dirname(script_dir), TEMPLATE_NAME),
        os.path.join(os.getcwd(), TEMPLATE_NAME),
    ]
    
    template_path = None
    for p in search_paths:
        if os.path.exists(p):
            template_path = p
            break
    
    if template_path:
        print(f"Loading template from: {template_path}")
        wb = openpyxl.load_workbook(template_path)
        ws = wb['1. Inputs']
        prefill_template(ws, data, ticker)
    else:
        print(f"Template '{TEMPLATE_NAME}' not found in:")
        for p in search_paths:
            print(f"  - {p}")
        print("Building standalone workbook with raw data instead...")
        wb = build_standalone_workbook(data, ticker)
    
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    return output_path


def write_row(ws, row_num: int, values: list, max_cols: int = 11):
    """Write values into template row, right-aligned (latest in col L=12).

    Takes up to max_cols values from the END of the list.
    Column B=2 is oldest, column L=12 is latest/TTM.
    """
    if not values:
        return
    n = min(len(values), max_cols)
    tail = values[-n:]
    for i, val in enumerate(tail):
        col = (max_cols - n) + i + 2  # right-align into B..L
        if 2 <= col <= 12 and val is not None:
            ws.cell(row=row_num, column=col, value=val)


def prefill_template(ws, data: dict, ticker: str):
    """Fill in the template's input sheet with fetched data."""
    
    # Company info
    ws['B2'] = data.get('company_name', ticker)
    ws['B3'] = ticker
    ws['B4'] = data.get('current_price', 0)
    ws['B5'] = round(data.get('shares', 0), 2)
    ws['B6'] = ''  # Sector - would need additional scraping
    ws['B7'] = 'STANDARD'  # Default
    
    # Income Statement data (rows 11-27)
    inc = data.get('income_statement', {})
    
    # Get year headers
    year_headers, _ = find_metric(inc, ['Sales', 'Revenue', 'Total Income', 'Net Interest Income', 'Interest Earned'])
    if not year_headers:
        year_headers, _ = find_metric(inc, ['Operating Profit', 'Net Profit', 'EPS'])
    
    # Determine how many years we have (up to 11: 10 years + TTM)
    n_years = min(len(year_headers), 11)
    
    # Write year labels (row 11) — right-aligned so latest lands in col L
    for i in range(n_years):
        col = (11 - n_years) + i + 2
        if 2 <= col <= 12:
            ws.cell(row=11, column=col, value=year_headers[-(n_years-i)])
    
    # Map metrics to rows
    income_mappings = [
        (12, ['Sales', 'Revenue', 'Total Income from Operations']),
        (13, ['Operating Profit']),
        (14, ['Depreciation']),
        (15, ['Interest']),
        (16, ['Profit before tax']),
        (17, ['Tax']),
        (18, ['Net Profit']),
        (19, ['EPS']),
        # DPS handled separately below — 'Dividend Payout' matches the payout
        # RATIO (%), not per-share dividend amounts.
    ]
    
    for row_num, keywords in income_mappings:
        _, values = find_metric(inc, keywords)
        if values:
            write_row(ws, row_num, values)
    
    # DPS: Screener's web table has "Dividend Payout %" (a ratio, not DPS).
    # Look for "Dividend Per Share" first; if absent, leave row 20 empty
    # for the user to fill manually.
    _, dps_values = find_metric(inc, ['Dividend Per Share'])
    if dps_values:
        write_row(ws, 20, dps_values)
    
    # Balance Sheet
    bs = data.get('balance_sheet', {})
    bs_mappings = [
        (31, ['Total Assets'], None),
        (32, ['Total Liabilities'], None),
        (33, ['Total Shareholders Funds', 'Shareholders Fund', "Shareholders' Equity"],
             ['equity share capital']),  # exclude the face-value-only row
        (34, ['Borrowings', 'Total Debt'], None),
        (35, ['Cash', 'Cash Equivalents'], None),
        (36, ['Goodwill', 'Intangibles'], None),
        (37, ['Reserves'], None),
    ]
    
    for row_num, keywords, exclusions in bs_mappings:
        _, values = find_metric(bs, keywords, exclude=exclusions)
        if values:
            write_row(ws, row_num, values)
    
    # Fallback for row 33 (Shareholders' Equity): if direct match failed,
    # compute as Equity Share Capital + Reserves
    if not any(ws.cell(row=33, column=c).value for c in range(2, 13)):
        _, sc_vals = find_metric(bs, ['Equity Share Capital', 'Share Capital'])
        _, res_vals = find_metric(bs, ['Reserves'])
        if sc_vals and res_vals:
            n = min(len(sc_vals), len(res_vals))
            combined = [(sc_vals[-(n-i)] or 0) + (res_vals[-(n-i)] or 0)
                        for i in range(n)]
            write_row(ws, 33, combined)
            print("  → Equity computed as Share Capital + Reserves (fallback)")
    
    # Cash Flow
    cf = data.get('cash_flow', {})
    cf_mappings = [
        (53, ['Cash from Operating', 'Operating Activity'], None),
        (54, ['Fixed Assets Purchased', 'Purchase of Fixed Assets', 'Capital Expenditure'], None),
        (55, ['Cash from Investing', 'Investing Activity'], None),
        (56, ['Cash from Financing', 'Financing Activity'], None),
    ]
    
    for row_num, keywords, exclusions in cf_mappings:
        _, values = find_metric(cf, keywords, exclude=exclusions)
        if values:
            write_row(ws, row_num, values)
    
    # Ratios
    ratios = data.get('ratios', {})
    
    # Try to get BVPS from ratios or balance sheet
    _, bvps_values = find_metric(ratios, ['Book Value'])
    if not bvps_values:
        # Compute from total shareholders' equity / shares
        # First try the direct total equity row
        _, equity_vals = find_metric(bs, ['Total Shareholders Funds', 'Shareholders Fund'],
                                     exclude=['equity share capital'])
        if not equity_vals:
            # Fallback: sum Share Capital + Reserves
            _, sc_vals = find_metric(bs, ['Equity Share Capital', 'Share Capital'])
            _, res_vals = find_metric(bs, ['Reserves'])
            if sc_vals and res_vals:
                n = min(len(sc_vals), len(res_vals))
                equity_vals = [(sc_vals[i] or 0) + (res_vals[i] or 0) for i in range(n)]
        if equity_vals and data.get('shares', 0) > 0:
            bvps_values = [v / data['shares'] if v else None for v in equity_vals]
    
    if bvps_values:
        write_row(ws, 38, bvps_values)
    
    # DPS from ratios — look for actual "Dividend Per Share", NOT "Dividend Payout"
    # which is the payout ratio (%) and would corrupt downstream calculations.
    _, dps_values = find_metric(ratios, ['Dividend Per Share'])
    if dps_values:
        write_row(ws, 20, dps_values)
    
    # Additional inputs — PE to row 67, P/B to row 68
    if data.get('current_price', 0) > 0:
        # Current PE
        latest_eps = None
        for c in range(12, 1, -1):
            v = ws.cell(row=19, column=c).value
            if v and v > 0:
                latest_eps = v
                break
        if latest_eps:
            ws.cell(row=67, column=2, value=round(data['current_price'] / latest_eps, 2))
        # Current P/B
        latest_bvps = None
        for c in range(12, 1, -1):
            v = ws.cell(row=38, column=c).value
            if v and v > 0:
                latest_bvps = v
                break
        if latest_bvps:
            ws.cell(row=68, column=2, value=round(data['current_price'] / latest_bvps, 2))
    
    print(f"Pre-filled {n_years} years of data for {data.get('company_name', ticker)}")


def build_standalone_workbook(data: dict, ticker: str):
    """Build a simple standalone workbook when template is not available."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Data"
    
    ws['A1'] = f"Financial Data for {data.get('company_name', ticker)}"
    ws['A1'].font = Font(size=14, bold=True)
    
    ws['A3'] = "Current Price:"
    ws['B3'] = data.get('current_price', 0)
    ws['A4'] = "Market Cap (Cr):"
    ws['B4'] = data.get('market_cap', 0)
    ws['A5'] = "PE Ratio:"
    ws['B5'] = data.get('pe_ratio', 0)
    ws['A6'] = "Shares (Cr):"
    ws['B6'] = round(data.get('shares', 0), 2)
    
    row = 8
    for section_name, section_key in [
        ("INCOME STATEMENT", "income_statement"),
        ("BALANCE SHEET", "balance_sheet"),
        ("CASH FLOW", "cash_flow"),
        ("RATIOS", "ratios"),
    ]:
        section = data.get(section_key, {})
        if not section:
            continue
        
        ws.cell(row=row, column=1, value=section_name).font = Font(bold=True, size=12)
        row += 1
        
        header_written = False
        for metric_name, metric_data in section.items():
            headers = metric_data.get('headers', [])
            values = metric_data.get('values', [])
            
            if not header_written and headers:
                ws.cell(row=row, column=1, value="Metric").font = Font(bold=True)
                for i, h in enumerate(headers):
                    ws.cell(row=row, column=i+2, value=h).font = Font(bold=True)
                row += 1
                header_written = True
            
            ws.cell(row=row, column=1, value=metric_name)
            for i, v in enumerate(values):
                if v is not None:
                    ws.cell(row=row, column=i+2, value=v)
            row += 1
        
        row += 1
    
    return wb


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage: python fetch_and_build_analysis.py TICKER")
        print("Example: python fetch_and_build_analysis.py HDFCBANK")
        print("\nThis script fetches financial data from Screener.in (free)")
        print("and creates a pre-filled Buffett Analysis Excel workbook.")
        print("\nIf 'Buffett_Analysis_Template.xlsx' is in the same directory,")
        print("it will populate that template. Otherwise, it creates a raw data sheet.")
        sys.exit(1)
    
    ticker = sys.argv[1].upper().strip()
    print(f"\n{'='*60}")
    print(f"  BUFFETT ANALYSIS DATA FETCHER")
    print(f"  Ticker: {ticker}")
    print(f"{'='*60}\n")
    
    # Fetch data
    print("Step 1: Fetching financial data from Screener.in...")
    data = fetch_screener_data(ticker)
    
    if not data or not data.get('income_statement'):
        print(f"\nWARNING: Could not fetch data for '{ticker}'.")
        print("Possible reasons:")
        print("  - Ticker not found on Screener.in")
        print("  - Network/rate limiting issue")
        print("  - Try the exact ticker used on Screener.in (e.g., HDFCBANK, TCS, ITC)")
        print("\nYou can still use the template manually.")
        sys.exit(1)
    
    print(f"  Company: {data.get('company_name', 'N/A')}")
    print(f"  CMP: ₹{data.get('current_price', 'N/A')}")
    print(f"  Market Cap: ₹{data.get('market_cap', 'N/A')} Cr")
    print(f"  PE: {data.get('pe_ratio', 'N/A')}")
    print(f"  Income Statement: {len(data.get('income_statement', {}))} metrics")
    print(f"  Balance Sheet: {len(data.get('balance_sheet', {}))} metrics")
    print(f"  Cash Flow: {len(data.get('cash_flow', {}))} metrics")
    
    # Build Excel
    output_file = f"{ticker}_Buffett_Analysis.xlsx"
    print(f"\nStep 2: Building Excel workbook...")
    build_prefilled_excel(ticker, data, output_file)
    
    print(f"\n{'='*60}")
    print(f"  DONE! Output: {output_file}")
    print(f"{'='*60}")
    print(f"\nNext steps:")
    print(f"  1. Open the Excel file")
    print(f"  2. Review and verify the auto-filled data")
    print(f"  3. Fill in any missing inputs (yellow cells)")
    print(f"  4. Check computed sheets for calculations")
    print(f"  5. Fill in DCF assumptions (Sheet 9)")
    print(f"  6. Score the final scorecard (Sheet 17)")


if __name__ == '__main__':
    main()
